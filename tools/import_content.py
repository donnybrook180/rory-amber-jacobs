#!/usr/bin/env python3
"""
Import & normalize the language master spreadsheet into app-ready CDN content.

Reads the Excel master (e.g. resources/Top_5000_Engelse_Woorden_Meertalig_Master.xlsx),
skips the info sheet, normalizes every row, validates, and writes:

    content/<lang>/words.<lang>.json   # normalized word records
    content/<lang>/manifest.json       # version, native langs, counts

Usage:
    python tools/import_content.py \
        --source resources/Top_5000_Engelse_Woorden_Meertalig_Master.xlsx \
        --target-lang en \
        --out content/english

Re-running is deterministic: keys are derived from Rank (en_0001 ...), so progress
data on devices stays valid across re-imports. The manifest `version` only bumps when
the produced content actually changes (content hash), so the app skips no-op downloads.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# --- configuration -----------------------------------------------------------

DATA_SHEET = "Top 5000 Master"          # the sheet that holds word rows
EXAMPLES_SHEET = "Examples"             # optional sheet of example sentences (Inc 8c)
INFO_SHEET_HINTS = ("overzicht", "info")  # sheets to ignore (substring match)

# Examples sheet column positions (0-based): Rank | Example (target) | native cols…
EX_COL_RANK = 0
EX_COL_TEXT = 1
EX_FIRST_LANG_COL = 2

# Fixed columns (0-based positions in the master sheet).
COL_RANK = 0
COL_WORD = 1
COL_POS = 2
FIRST_LANG_COL = 3                      # native-language columns start here

# Map a native-language header -> ISO code. Matched by the parenthesised code in
# the header, e.g. "Español (ES)" -> "es". Falls back to that code lowercased.
HEADER_CODE_RE = re.compile(r"\(([A-Za-z]{2,3})\)\s*$")

# Localized Part-of-Speech -> neutral enum. Unknown values map to "other" (warned).
POS_MAP = {
    "werkwoord": "verb",
    "bijvoeglijk naamwoord": "adjective",
    "zelfstandig naamwoord": "noun",
    "voorzetsel": "preposition",
    "lidwoord/bepaling": "article",
    "voegwoord": "conjunction",
    "voornaamwoord": "pronoun",
    "conjunction/pronoun": "pronoun",
}

# Rank bands -> difficulty level.
def level_for_rank(rank: int) -> str:
    if rank <= 1000:
        return "beginner"
    if rank <= 3000:
        return "intermediate"
    return "advanced"


# Learner Packs: group words into chunks of 10 by rank. The UX shows these as
# "Woorden 1261–1270". pack_id is 1-based; pack_label is the rank range.
PACK_SIZE = 10

def pack_for_rank(rank: int) -> dict:
    pack_id = (rank - 1) // PACK_SIZE + 1
    start = (pack_id - 1) * PACK_SIZE + 1
    end = start + PACK_SIZE - 1
    return {"id": pack_id, "label": f"{start}-{end}"}


# --- helpers -----------------------------------------------------------------

def split_translations(cell) -> list[str]:
    """A cell may hold several comma-separated options: 'ser, estar' -> [ser, estar]."""
    if cell is None:
        return []
    parts = [p.strip() for p in str(cell).split(",")]
    return [p for p in parts if p]


def header_to_code(header: str) -> str:
    m = HEADER_CODE_RE.search(header or "")
    if m:
        return m.group(1).lower()
    # Fallback: slug of the header.
    return re.sub(r"[^a-z]", "", (header or "").lower())[:3] or "xx"


def content_hash(words: list[dict]) -> str:
    blob = json.dumps(words, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()[:12]


# --- core --------------------------------------------------------------------

def read_examples(wb, target_lang: str, warnings: list[str]) -> dict[str, list[dict]]:
    """Parse the optional `Examples` sheet → {word key: [example, …]} (Inc 8c).

    One row per sentence, foreign-keyed by Rank (repeat the Rank for multiple
    sentences). Native columns use the same `(XX)` header convention as the word
    sheet; each cell holds a single translated sentence (no comma-splitting).
    Returns an empty map when the sheet is absent.
    """
    if EXAMPLES_SHEET not in wb.sheetnames:
        return {}

    rows = wb[EXAMPLES_SHEET].iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return {}

    lang_cols = []  # (col_index, iso_code)
    for idx in range(EX_FIRST_LANG_COL, len(header)):
        if header[idx]:
            lang_cols.append((idx, header_to_code(header[idx])))

    by_key: dict[str, list[dict]] = {}
    for line_no, row in enumerate(rows, start=2):
        if row is None or row[EX_COL_TEXT] in (None, ""):
            continue
        try:
            rank = int(row[EX_COL_RANK])
        except (TypeError, ValueError):
            warnings.append(f"examples row {line_no}: non-numeric rank "
                            f"{row[EX_COL_RANK]!r}, skipped")
            continue

        key = f"{target_lang}_{rank:04d}"
        translations = {}
        for col_idx, code in lang_cols:
            val = row[col_idx] if col_idx < len(row) else None
            if val not in (None, ""):
                translations[code] = str(val).strip()

        by_key.setdefault(key, []).append({
            "text": str(row[EX_COL_TEXT]).strip(),
            "translations": translations,
            "audio": None,
        })
    return by_key


def import_workbook(source: Path, target_lang: str, audio_dir: Path | None = None):
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)

    sheet_name = DATA_SHEET if DATA_SHEET in wb.sheetnames else None
    if sheet_name is None:
        # Pick the first sheet that is not an info sheet.
        for name in wb.sheetnames:
            if not any(h in name.lower() for h in INFO_SHEET_HINTS):
                sheet_name = name
                break
    if sheet_name is None:
        sys.exit(f"No data sheet found in {source.name}; sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)
    lang_cols = []  # (col_index, iso_code)
    for idx in range(FIRST_LANG_COL, len(header)):
        if header[idx]:
            lang_cols.append((idx, header_to_code(header[idx])))
    native_langs = [code for _, code in lang_cols]

    words: list[dict] = []
    warnings: list[str] = []
    seen_keys: set[str] = set()

    for line_no, row in enumerate(rows, start=2):
        if row is None or row[COL_WORD] in (None, ""):
            continue
        rank = row[COL_RANK]
        try:
            rank = int(rank)
        except (TypeError, ValueError):
            warnings.append(f"row {line_no}: non-numeric rank {rank!r}, skipped")
            continue

        key = f"{target_lang}_{rank:04d}"
        if key in seen_keys:
            warnings.append(f"row {line_no}: duplicate rank {rank} -> key {key}, skipped")
            continue
        seen_keys.add(key)

        raw_pos = (row[COL_POS] or "").strip().lower()
        pos = POS_MAP.get(raw_pos)
        if pos is None:
            pos = "other"
            if raw_pos:
                warnings.append(f"row {line_no}: unknown POS {raw_pos!r} -> 'other'")

        translations = {}
        for col_idx, code in lang_cols:
            vals = split_translations(row[col_idx] if col_idx < len(row) else None)
            if not vals:
                warnings.append(f"row {line_no} ({key}): missing '{code}' translation")
            translations[code] = vals

        # Audio is opt-in: set <key>.mp3 only when the file actually ships in
        # the audio/ folder, so the app hides the play button otherwise.
        audio = None
        if audio_dir is not None and (audio_dir / f"{key}.mp3").exists():
            audio = f"{key}.mp3"

        words.append({
            "key": key,
            "rank": rank,
            "target": str(row[COL_WORD]).strip(),
            "pos": pos,
            "level": level_for_rank(rank),
            "pack": pack_for_rank(rank),
            "translations": translations,
            "audio": audio,
            "examples": [],
        })

    # Inc 8c: attach example sentences (optional `Examples` sheet), keyed by word key.
    examples_by_key = read_examples(wb, target_lang, warnings)
    if examples_by_key:
        by_key = {w["key"]: w for w in words}
        for key, examples in examples_by_key.items():
            target_word = by_key.get(key)
            if target_word is None:
                warnings.append(f"examples: rank for key {key} not in word sheet, "
                                f"{len(examples)} sentence(s) dropped")
                continue
            target_word["examples"] = examples

    words.sort(key=lambda w: w["rank"])
    return words, native_langs, warnings


def write_output(out_dir: Path, target_lang: str, words: list[dict],
                 native_langs: list[str], min_app_version: str):
    out_dir.mkdir(parents=True, exist_ok=True)

    words_path = out_dir / f"words.{target_lang}.json"
    words_path.write_text(
        json.dumps(words, ensure_ascii=False, indent=2), encoding="utf-8")

    new_hash = content_hash(words)
    manifest_path = out_dir / "manifest.json"
    prev_version = 0
    if manifest_path.exists():
        try:
            prev = json.loads(manifest_path.read_text(encoding="utf-8"))
            prev_version = prev.get("version", 0)
            if prev.get("contentHash") == new_hash:
                # No content change -> keep version, still rewrite words file is fine.
                print(f"  content unchanged (hash {new_hash}); version stays {prev_version}")
                return words_path, manifest_path, prev_version
        except (json.JSONDecodeError, OSError):
            pass

    manifest = {
        "version": prev_version + 1,
        "targetLang": target_lang,
        "nativeLangs": native_langs,
        "wordCount": len(words),
        "audioCount": sum(1 for w in words if w["audio"]),
        "exampleCount": sum(1 for w in words if w["examples"]),
        "contentHash": new_hash,
        "minAppVersion": min_app_version,
        "files": {"words": words_path.name},
    }
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return words_path, manifest_path, manifest["version"]


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--source", required=True, type=Path, help="path to the .xlsx master")
    ap.add_argument("--target-lang", required=True, help="ISO code of the language being taught, e.g. en")
    ap.add_argument("--out", required=True, type=Path, help="output content dir, e.g. content/english")
    ap.add_argument("--audio-dir", type=Path, default=None,
                    help="folder of <key>.mp3 files (default: <out>/audio)")
    ap.add_argument("--min-app-version", default="1.0.0")
    ap.add_argument("--strict", action="store_true",
                    help="exit non-zero if there are any validation warnings")
    args = ap.parse_args()

    if not args.source.exists():
        sys.exit(f"Source not found: {args.source}")

    audio_dir = args.audio_dir or (args.out / "audio")
    words, native_langs, warnings = import_workbook(
        args.source, args.target_lang, audio_dir if audio_dir.exists() else None)

    if not words:
        sys.exit("No word rows imported — check the sheet name and layout.")

    words_path, manifest_path, version = write_output(
        args.out, args.target_lang, words, native_langs, args.min_app_version)

    audio_count = sum(1 for w in words if w["audio"])
    example_count = sum(1 for w in words if w["examples"])
    print(f"Imported {len(words)} words; native langs ({len(native_langs)}): "
          f"{', '.join(native_langs)}; audio files: {audio_count}; "
          f"words with examples: {example_count}")
    print(f"  -> {words_path}")
    print(f"  -> {manifest_path} (version {version})")

    if warnings:
        shown = warnings[:20]
        print(f"\n{len(warnings)} validation warning(s):")
        for w in shown:
            print(f"  ! {w}")
        if len(warnings) > len(shown):
            print(f"  ... and {len(warnings) - len(shown)} more")
        if args.strict:
            sys.exit(1)


if __name__ == "__main__":
    main()
